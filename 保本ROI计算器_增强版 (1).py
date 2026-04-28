#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保本ROI计算器与推广盈亏计算器 - 增强版 Pro
作者：Grok x 用户定制
功能：保本价计算、ROI分析、推广盈亏、批量计算、图表可视化、Excel导出
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import ttkbootstrap as ttkb
from ttkbootstrap.constants import *
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json
import os
import sys

# 默认费率配置（中国主流电商平台）
PLATFORM_PRESETS = {
    "淘宝/天猫": {"commission": 6.8, "tax": 6.0, "other": 0.5},
    "拼多多": {"commission": 5.0, "tax": 6.0, "other": 0.3},
    "抖音/快手": {"commission": 8.0, "tax": 6.0, "other": 1.0},
    "京东": {"commission": 8.0, "tax": 6.0, "other": 0.8},
    "小红书": {"commission": 10.0, "tax": 6.0, "other": 1.5},
    "自定义": {"commission": 6.0, "tax": 6.0, "other": 0.0}
}


class Tooltip:
    """鼠标悬停提示工具（支持 ttkbootstrap 样式）"""
    def __init__(self, widget, text, delay=400):
        self.widget = widget
        self.text = text
        self.delay = delay
        self.tooltip = None
        self.id = None
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hide()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(self.delay, self.show)

    def unschedule(self):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None

    def show(self):
        if self.tooltip or not self.widget.winfo_exists():
            return
        # 计算位置（在控件下方）
        x = self.widget.winfo_rootx() + 15
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        # 使用 ttkbootstrap 样式
        label = ttkb.Label(
            self.tooltip, 
            text=self.text, 
            bootstyle="info", 
            padding=10, 
            wraplength=320, 
            justify="left",
            font=("Microsoft YaHei", 10)
        )
        label.pack()

    def hide(self):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None


class ROICalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("保本ROI计算器 Pro - 增强版")
        self.root.geometry("1100x750")
        self.style = ttkb.Style(theme="cosmo")  # 现代主题，可切换 dark/superhero

        # 设置专属屎哥图标（支持 PyInstaller --onefile 打包）
        try:
            if hasattr(sys, '_MEIPASS'):
                # 打包后的 exe 内部路径
                icon_path = os.path.join(sys._MEIPASS, "屎哥.ico")
            else:
                icon_path = "屎哥.ico"
            self.root.iconbitmap(icon_path)
        except Exception:
            pass  # 跨平台或无图标时静默

        # 历史记录
        self.history = []
        self.current_platform = "淘宝/天猫"

        self.create_widgets()
        self.load_defaults()

    def create_widgets(self):
        # 顶部工具栏
        toolbar = ttk.Frame(self.root, padding=10)
        toolbar.pack(fill=X)

        # 屎哥 红色图标（左上角）- 使用 poop emoji 做可爱图标
        屎哥_frame = ttkb.Frame(toolbar, bootstyle="danger", padding=5)
        屎哥_label = ttkb.Label(
            屎哥_frame, 
            text="💩", 
            font=("Segoe UI Emoji", 20), 
            foreground="white"
        )
        屎哥_label.pack()
        屎哥_frame.pack(side=LEFT, padx=8)

        ttkb.Label(toolbar, text="🛠️ 保本ROI计算器 Pro", font=("Microsoft YaHei", 16, "bold")).pack(side=LEFT)
        
        # 主题切换
        theme_btn = ttkb.Button(toolbar, text="🌙 切换深色模式", command=self.toggle_theme, bootstyle="secondary")
        theme_btn.pack(side=RIGHT, padx=5)

        # 帮助按钮
        help_btn = ttkb.Button(toolbar, text="❓ 使用说明", command=self.show_help, bootstyle="info")
        help_btn.pack(side=RIGHT, padx=5)

        # 主 Notebook (多标签页)
        self.notebook = ttk.Notebook(self.root, padding=10)
        self.notebook.pack(fill=BOTH, expand=True, padx=10, pady=5)

        # Tab 1: 保本价计算
        self.tab1 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab1, text="📊 保本价计算")
        self.create_breakeven_tab()

        # Tab 2: ROI & 推广盈亏
        self.tab2 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab2, text="💰 ROI & 推广盈亏")
        self.create_roi_tab()

        # Tab 3: 批量计算
        self.tab3 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab3, text="📋 批量计算 & 对比")
        self.create_batch_tab()

        # Tab 4: 设置
        self.tab4 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab4, text="⚙️ 高级设置")
        self.create_settings_tab()

        # 底部状态栏
        status = ttk.Frame(self.root, padding=5)
        status.pack(fill=X, side=BOTTOM)
        self.status_var = tk.StringVar(value="就绪 | 支持多平台费率预设 | 可导出Excel报告")
        ttkb.Label(status, textvariable=self.status_var, bootstyle="secondary").pack(side=LEFT)
        ttkb.Label(status, text="v2.0 Enhanced | 2026", bootstyle="secondary").pack(side=RIGHT)

    def create_breakeven_tab(self):
        frame = ttk.LabelFrame(self.tab1, text="保本价计算器", padding=15)
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # 输入区
        input_frame = ttk.Frame(frame)
        input_frame.pack(fill=X, pady=10)

        # 左侧输入
        left = ttk.Frame(input_frame)
        left.pack(side=LEFT, fill=X, expand=True, padx=20)

        fields = [
            ("商品成本 (元)", "cost", 50.0),
            ("目标利润率 (%)", "profit_rate", 15.0),
            ("预计销量 (件)", "expected_sales", 100),
            ("固定费用总额 (元)", "fixed_cost", 2000.0),
        ]

        self.breakeven_vars = {}
        tooltip_texts = [
            "商品的采购成本或生产成本（元/件），不包含任何推广费用",
            "你希望每卖出一件商品获得的纯利润占售价的百分比（建议 10%-30%）",
            "预估这个商品能卖出去多少件（用于把固定费用分摊到每件上）",
            "本次推广的总固定投入（如广告费、装修、人工、道具等总额）"
        ]
        for i, (label, key, default) in enumerate(fields):
            ttkb.Label(left, text=label, font=("Microsoft YaHei", 11)).grid(row=i, column=0, sticky=W, pady=8)
            var = tk.DoubleVar(value=default)
            entry = ttkb.Entry(left, textvariable=var, width=18, bootstyle="success")
            entry.grid(row=i, column=1, pady=8, padx=10)
            self.breakeven_vars[key] = var
            # 添加悬停提示
            Tooltip(entry, tooltip_texts[i])

        # 右侧平台选择
        right = ttk.Frame(input_frame)
        right.pack(side=RIGHT, fill=X, expand=True, padx=20)

        ttkb.Label(right, text="平台费率预设", font=("Microsoft YaHei", 11, "bold")).pack(anchor=W, pady=5)
        self.platform_var = tk.StringVar(value="淘宝/天猫")
        self.platform_combo = ttkb.Combobox(right, textvariable=self.platform_var, values=list(PLATFORM_PRESETS.keys()), 
                                        state="readonly", width=18, bootstyle="primary")
        self.platform_combo.pack(anchor=W, pady=5)
        self.platform_combo.bind("<<ComboboxSelected>>", self.update_platform_rates)
        Tooltip(platform_combo, "选择主流电商平台后会自动填充对应的佣金率、税率和其他费用率\n\n淘宝/天猫 ≈ 6.8%+6%\n拼多多 ≈ 5%+6%\n抖音 ≈ 8%+6%")

        # 动态费率显示
        rate_frame = ttk.Frame(right)
        rate_frame.pack(anchor=W, pady=10)
        self.comm_label = ttkb.Label(rate_frame, text="佣金率: 6.8%")
        self.comm_label.pack(anchor=W)
        self.tax_label = ttkb.Label(rate_frame, text="税率: 6.0%")
        self.tax_label.pack(anchor=W)
        self.other_label = ttkb.Label(rate_frame, text="其他费用率: 0.5%")
        self.other_label.pack(anchor=W)

        # 计算按钮
        calc_btn = ttkb.Button(frame, text="🚀 计算保本价", command=self.calculate_breakeven, 
                               bootstyle="success", width=25)
        calc_btn.pack(pady=15)
        Tooltip(calc_btn, "点击后立即计算保本价、单件毛利、净利润、成本结构饼图等\n\n会根据你选择的平台费率自动扣除佣金+税费+其他费用")

        # 结果区
        result_frame = ttk.LabelFrame(frame, text="计算结果", padding=15)
        result_frame.pack(fill=X, pady=10)

        self.breakeven_result = tk.StringVar(value="请点击上方按钮计算...")
        ttkb.Label(result_frame, textvariable=self.breakeven_result, font=("Microsoft YaHei", 12), 
                   wraplength=600, justify=LEFT).pack(anchor=W)

        # 图表区
        chart_frame = ttk.Frame(frame)
        chart_frame.pack(fill=BOTH, expand=True, pady=10)
        self.breakeven_fig = plt.Figure(figsize=(6, 3.5), dpi=100)
        self.breakeven_canvas = FigureCanvasTkAgg(self.breakeven_fig, master=chart_frame)
        self.breakeven_canvas.get_tk_widget().pack(fill=BOTH, expand=True)

    def update_platform_rates(self, event=None):
        platform = self.platform_var.get()
        rates = PLATFORM_PRESETS.get(platform, PLATFORM_PRESETS["自定义"])
        self.comm_label.config(text=f"佣金率: {rates['commission']}%")
        self.tax_label.config(text=f"税率: {rates['tax']}%")
        self.other_label.config(text=f"其他费用率: {rates['other']}%")

    def calculate_breakeven(self):
        try:
            cost = self.breakeven_vars["cost"].get()
            profit_rate = self.breakeven_vars["profit_rate"].get() / 100
            sales = self.breakeven_vars["expected_sales"].get()
            fixed = self.breakeven_vars["fixed_cost"].get()

            platform = self.platform_var.get()
            rates = PLATFORM_PRESETS.get(platform, PLATFORM_PRESETS["自定义"])
            comm = rates["commission"] / 100
            tax = rates["tax"] / 100
            other_rate = rates["other"] / 100

            # 保本价公式（考虑所有费用后净利润为0）
            # 简化模型：保本价 = (成本 + 固定费用/销量) / (1 - 佣金 - 税 - 其他率 - 目标利润率)
            denominator = 1 - comm - tax - other_rate - profit_rate
            if denominator <= 0:
                messagebox.showerror("错误", "费率 + 利润率总和不能 ≥ 100%，请调整！")
                return

            breakeven_price = (cost + fixed / sales) / denominator
            gross_profit = breakeven_price * (1 - comm - tax) - cost
            net_profit_per_unit = gross_profit - (fixed / sales) - other_rate * breakeven_price

            result_text = (
                f"💰【单件毛利润】 ¥{gross_profit:.2f}   ← 核心指标！\n\n"
                f"✅ 保本价: ¥{breakeven_price:.2f} /件\n"
                f"📈 目标利润率: {self.breakeven_vars['profit_rate'].get()}%\n"
                f"💵 单件毛利: ¥{gross_profit:.2f}\n"
                f"💰 单件净利润: ¥{net_profit_per_unit:.2f}\n"
                f"📊 总固定成本分摊: ¥{fixed/sales:.2f}/件\n\n"
                f"平台: {platform} | 佣金 {rates['commission']}% + 税 {rates['tax']}% + 其他 {rates['other']}%"
            )
            self.breakeven_result.set(result_text)

            # 绘制成本结构饼图
            self.draw_breakeven_chart(cost, breakeven_price, comm, tax, other_rate, profit_rate, fixed/sales)

            self.status_var.set(f"保本价计算完成 | {datetime.now().strftime('%H:%M:%S')}")
            self.add_to_history("保本价计算", f"成本¥{cost} → 保本价¥{breakeven_price:.2f}")

        except Exception as e:
            messagebox.showerror("计算错误", str(e))

    def draw_breakeven_chart(self, cost, price, comm, tax, other, profit, fixed_per_unit):
        self.breakeven_fig.clear()
        ax = self.breakeven_fig.add_subplot(111)

        labels = ['商品成本', '佣金', '税费', '其他费用', '目标利润', '固定成本分摊']
        sizes = [
            cost,
            price * comm,
            price * tax,
            price * other,
            price * profit,
            fixed_per_unit
        ]
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD']
        explode = (0.05, 0, 0, 0, 0.05, 0)

        ax.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%',
               shadow=True, startangle=90, textprops={'fontsize': 9})
        ax.set_title(f'保本价 ¥{price:.2f} 成本结构占比', fontsize=11, fontweight='bold')
        ax.axis('equal')

        self.breakeven_fig.tight_layout()
        self.breakeven_canvas.draw()

    def create_roi_tab(self):
        frame = ttk.LabelFrame(self.tab2, text="ROI 与 推广盈亏计算器", padding=15)
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # 输入
        input_frame = ttk.Frame(frame)
        input_frame.pack(fill=X, pady=10)

        left = ttk.Frame(input_frame)
        left.pack(side=LEFT, fill=X, expand=True, padx=15)

        roi_fields = [
            ("销售额 (元)", "revenue", 50000.0),
            ("商品总成本 (元)", "total_cost", 28000.0),
            ("推广费用 (元)", "ad_spend", 8000.0),
            ("其他费用 (元)", "other_cost", 1500.0),
        ]

        self.roi_vars = {}
        roi_tooltip_texts = [
            "本次推广或销售的总收入（不含退款）",
            "所有卖出商品的总成本（进货价 × 销量）",
            "本次推广花费的广告费、达人费、直播间费用等",
            "除了推广费之外的其他杂费（如包装、运费、客服等）"
        ]
        for i, (label, key, default) in enumerate(roi_fields):
            ttkb.Label(left, text=label, font=("Microsoft YaHei", 11)).grid(row=i, column=0, sticky=W, pady=6)
            var = tk.DoubleVar(value=default)
            entry = ttkb.Entry(left, textvariable=var, width=18, bootstyle="warning")
            entry.grid(row=i, column=1, pady=6, padx=8)
            self.roi_vars[key] = var
            Tooltip(entry, roi_tooltip_texts[i])

        right = ttk.Frame(input_frame)
        right.pack(side=RIGHT, fill=X, expand=True, padx=15)

        ttkb.Label(right, text="费率设置（同保本价）", font=("Microsoft YaHei", 11, "bold")).pack(anchor=W)
        self.roi_platform_var = tk.StringVar(value="淘宝/天猫")
        self.roi_combo = ttkb.Combobox(right, textvariable=self.roi_platform_var, values=list(PLATFORM_PRESETS.keys()),
                                   state="readonly", width=18, bootstyle="warning")
        self.roi_combo.pack(anchor=W, pady=5)
        self.roi_combo.bind("<<ComboboxSelected>>", lambda e: self.update_roi_rates())
        Tooltip(roi_combo, "选择平台后自动带出费率，影响净利润和ROI计算\n\n建议根据实际店铺所在平台选择")

        self.roi_comm_label = ttkb.Label(right, text="佣金率: 6.8%")
        self.roi_comm_label.pack(anchor=W)
        self.roi_tax_label = ttkb.Label(right, text="税率: 6.0%")
        self.roi_tax_label.pack(anchor=W)

        # 按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=10)
        calc_roi_btn = ttkb.Button(btn_frame, text="📈 计算 ROI & 盈亏", command=self.calculate_roi, 
                    bootstyle="warning", width=22)
        calc_roi_btn.pack(side=LEFT, padx=10)
        Tooltip(calc_roi_btn, "计算毛利润、净利润、ROI、盈亏平衡推广费上限等\n\n结果会实时更新右下角瀑布图")

        draw_btn = ttkb.Button(btn_frame, text="📊 绘制趋势图", command=self.draw_roi_chart, 
                    bootstyle="info", width=18)
        draw_btn.pack(side=LEFT, padx=10)
        Tooltip(draw_btn, "根据当前输入数据绘制推广盈亏瀑布图\n\n直观展示销售额→成本→推广费→佣金→税费→净利润的流向")

        # 结果
        result_frame = ttk.LabelFrame(frame, text="计算结果", padding=15)
        result_frame.pack(fill=X, pady=10)

        self.roi_result = tk.StringVar(value="点击计算按钮查看详细盈亏分析...")
        ttkb.Label(result_frame, textvariable=self.roi_result, font=("Microsoft YaHei", 11), 
                   wraplength=650, justify=LEFT).pack(anchor=W)

        # 图表
        chart_frame = ttk.Frame(frame)
        chart_frame.pack(fill=BOTH, expand=True, pady=5)
        self.roi_fig = plt.Figure(figsize=(7, 3.2), dpi=100)
        self.roi_canvas = FigureCanvasTkAgg(self.roi_fig, master=chart_frame)
        self.roi_canvas.get_tk_widget().pack(fill=BOTH, expand=True)

    def update_roi_rates(self):
        platform = self.roi_platform_var.get()
        rates = PLATFORM_PRESETS.get(platform, PLATFORM_PRESETS["自定义"])
        self.roi_comm_label.config(text=f"佣金率: {rates['commission']}%")
        self.roi_tax_label.config(text=f"税率: {rates['tax']}%")

    def calculate_roi(self):
        try:
            revenue = self.roi_vars["revenue"].get()
            total_cost = self.roi_vars["total_cost"].get()
            ad_spend = self.roi_vars["ad_spend"].get()
            other = self.roi_vars["other_cost"].get()

            platform = self.roi_platform_var.get()
            rates = PLATFORM_PRESETS.get(platform, PLATFORM_PRESETS["自定义"])
            comm = rates["commission"] / 100
            tax = rates["tax"] / 100

            # 计算
            gross_profit = revenue - total_cost
            commission_fee = revenue * comm
            tax_fee = revenue * tax
            net_profit = gross_profit - ad_spend - other - commission_fee - tax_fee

            roi = (net_profit / ad_spend * 100) if ad_spend > 0 else 0
            profit_margin = (net_profit / revenue * 100) if revenue > 0 else 0

            # 盈亏平衡推广费（净利=0时的最大推广预算）
            max_ad_for_breakeven = gross_profit - other - commission_fee - tax_fee
            if max_ad_for_breakeven < 0:
                max_ad_for_breakeven = 0

            result = (
                f"💰【毛利润】 ¥{gross_profit:,.2f}   ← 核心指标！\n\n"
                f"📊 销售额: ¥{revenue:,.2f}   |   总成本: ¥{total_cost:,.2f}\n"
                f"💵 毛利润: ¥{gross_profit:,.2f}   |   净利润: ¥{net_profit:,.2f}\n"
                f"📈 ROI: {roi:.1f}%   |   销售净利率: {profit_margin:.1f}%\n"
                f"💸 佣金: ¥{commission_fee:,.2f}   |   税费: ¥{tax_fee:,.2f}\n"
                f"⚠️ 推广盈亏平衡点: 推广费 ≤ ¥{max_ad_for_breakeven:,.2f} 时仍可保本\n\n"
                f"结论: {'✅ 盈利' if net_profit > 0 else '❌ 亏损'} | 平台: {platform}"
            )
            self.roi_result.set(result)

            self.status_var.set(f"ROI计算完成 | 净利润 ¥{net_profit:,.2f}")
            self.add_to_history("ROI计算", f"销售额¥{revenue} → ROI {roi:.1f}%")

        except Exception as e:
            messagebox.showerror("计算错误", str(e))

    def draw_roi_chart(self):
        try:
            revenue = self.roi_vars["revenue"].get()
            total_cost = self.roi_vars["total_cost"].get()
            ad_spend = self.roi_vars["ad_spend"].get()
            other = self.roi_vars["other_cost"].get()

            platform = self.roi_platform_var.get()
            rates = PLATFORM_PRESETS.get(platform, PLATFORM_PRESETS["自定义"])
            comm = rates["commission"] / 100
            tax = rates["tax"] / 100

            commission_fee = revenue * comm
            tax_fee = revenue * tax
            net = revenue - total_cost - ad_spend - other - commission_fee - tax_fee

            self.roi_fig.clear()
            ax = self.roi_fig.add_subplot(111)

            categories = ['销售额', '成本', '推广费', '佣金', '税费', '其他', '净利润']
            values = [revenue, -total_cost, -ad_spend, -commission_fee, -tax_fee, -other, net]
            colors = ['#2ECC71' if v >= 0 else '#E74C3C' for v in values]

            bars = ax.barh(categories, values, color=colors, edgecolor='black', linewidth=0.5)
            ax.axvline(0, color='black', linewidth=1)
            ax.set_xlabel('金额 (元)', fontsize=10)
            ax.set_title('推广盈亏瀑布图分析', fontsize=12, fontweight='bold')
            ax.grid(axis='x', alpha=0.3)

            for bar, val in zip(bars, values):
                width = bar.get_width()
                ax.text(width + (50 if width > 0 else -150), bar.get_y() + bar.get_height()/2, 
                        f'¥{abs(val):,.0f}', va='center', fontsize=9)

            self.roi_fig.tight_layout()
            self.roi_canvas.draw()

        except Exception as e:
            messagebox.showerror("绘图错误", str(e))

    def create_batch_tab(self):
        frame = ttk.LabelFrame(self.tab3, text="批量计算与产品对比", padding=15)
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # 说明
        ttkb.Label(frame, text="支持添加多款商品，自动计算保本价、ROI、盈亏，一键导出Excel报告", 
                   font=("Microsoft YaHei", 10)).pack(anchor=W, pady=5)

        # 表格
        columns = ("商品名称", "成本", "售价", "推广费", "销量", "保本价", "ROI%", "净利润", "状态")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings", height=12, bootstyle="primary")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=95, anchor=CENTER)
        self.tree.pack(fill=BOTH, expand=True, pady=10)

        # 控制按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=X, pady=5)

        ttkb.Button(btn_frame, text="➕ 添加商品", command=self.add_batch_item, bootstyle="success").pack(side=LEFT, padx=5)
        ttkb.Button(btn_frame, text="🗑️ 删除选中", command=self.delete_batch_item, bootstyle="danger").pack(side=LEFT, padx=5)
        ttkb.Button(btn_frame, text="🔄 全部计算", command=self.calculate_batch, bootstyle="warning").pack(side=LEFT, padx=5)
        ttkb.Button(btn_frame, text="📤 导出Excel", command=self.export_to_excel, bootstyle="info").pack(side=LEFT, padx=5)
        ttkb.Button(btn_frame, text="🧹 清空列表", command=self.clear_batch, bootstyle="secondary").pack(side=LEFT, padx=5)

        # 汇总
        self.summary_var = tk.StringVar(value="总计: 0款商品 | 总净利润: ¥0.00 | 平均ROI: 0%")
        ttkb.Label(frame, textvariable=self.summary_var, font=("Microsoft YaHei", 11, "bold"), bootstyle="success").pack(anchor=W, pady=5)

    def add_batch_item(self):
        # 简单弹窗添加
        dialog = tk.Toplevel(self.root)
        dialog.title("添加商品")
        dialog.geometry("350x280")

        entries = {}
        labels = ["商品名称", "成本(元)", "售价(元)", "推广费(元)", "销量(件)"]
        defaults = ["新品A", "45", "89", "12000", "300"]

        for i, (lab, defv) in enumerate(zip(labels, defaults)):
            ttkb.Label(dialog, text=lab).grid(row=i, column=0, padx=10, pady=8, sticky=W)
            var = tk.StringVar(value=defv)
            ttkb.Entry(dialog, textvariable=var, width=20).grid(row=i, column=1, padx=10, pady=8)
            entries[lab] = var

        def add():
            try:
                name = entries["商品名称"].get()
                cost = float(entries["成本(元)"].get())
                price = float(entries["售价(元)"].get())
                ad = float(entries["推广费(元)"].get())
                sales = int(entries["销量(件)"].get())
                self.tree.insert("", "end", values=(name, cost, price, ad, sales, "-", "-", "-", "-"))
                dialog.destroy()
            except:
                messagebox.showerror("输入错误", "请检查数值输入是否正确")

        ttkb.Button(dialog, text="确认添加", command=add, bootstyle="success").grid(row=5, column=0, columnspan=2, pady=15)

    def delete_batch_item(self):
        selected = self.tree.selection()
        if selected:
            self.tree.delete(selected[0])

    def calculate_batch(self):
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            name, cost, price, ad, sales = values[0], float(values[1]), float(values[2]), float(values[3]), int(values[4])

            # 简单计算（使用当前平台费率）
            platform = self.platform_var.get() if hasattr(self, 'platform_var') else "淘宝/天猫"
            rates = PLATFORM_PRESETS.get(platform, PLATFORM_PRESETS["自定义"])
            comm = rates["commission"] / 100
            tax = rates["tax"] / 100

            gross = price * sales - cost * sales
            total_cost = cost * sales + ad + (price * sales * comm) + (price * sales * tax)
            net = gross - ad - (price * sales * comm) - (price * sales * tax)
            roi = (net / ad * 100) if ad > 0 else 0
            breakeven = (cost + (ad + 2000) / sales) / (1 - comm - tax) if sales > 0 else 0  # 简化

            status = "✅ 盈利" if net > 0 else "❌ 亏损"
            self.tree.item(item, values=(name, cost, price, ad, sales, f"¥{breakeven:.1f}", f"{roi:.1f}%", f"¥{net:,.0f}", status))

        # 更新汇总
        total_net = 0
        total_roi = 0
        count = 0
        for item in self.tree.get_children():
            vals = self.tree.item(item, "values")
            try:
                net_str = vals[7].replace("¥", "").replace(",", "")
                total_net += float(net_str)
                roi_str = vals[6].replace("%", "")
                total_roi += float(roi_str)
                count += 1
            except:
                pass

        avg_roi = total_roi / count if count > 0 else 0
        self.summary_var.set(f"总计: {count}款商品 | 总净利润: ¥{total_net:,.2f} | 平均ROI: {avg_roi:.1f}%")

    def clear_batch(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.summary_var.set("总计: 0款商品 | 总净利润: ¥0.00 | 平均ROI: 0%")

    def export_to_excel(self):
        if not self.tree.get_children():
            messagebox.showwarning("提示", "请先添加并计算商品！")
            return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 filetypes=[("Excel文件", "*.xlsx")],
                                                 title="保存报告")
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "推广盈亏报告"

        # 标题
        ws.merge_cells('A1:I1')
        ws['A1'] = f"保本ROI计算器 Pro - 推广盈亏分析报告 ({datetime.now().strftime('%Y-%m-%d %H:%M')})"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")

        # 表头
        headers = ["商品名称", "成本(元)", "售价(元)", "推广费(元)", "销量(件)", "保本价(元)", "ROI(%)", "净利润(元)", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 数据
        for row_idx, item in enumerate(self.tree.get_children(), 4):
            values = self.tree.item(item, "values")
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center")
                if col_idx == 9:  # 状态列
                    if "盈利" in str(val):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 汇总
        last_row = 4 + len(self.tree.get_children())
        ws.cell(row=last_row + 1, column=1, value="汇总").font = Font(bold=True)
        ws.cell(row=last_row + 1, column=8, value=self.summary_var.get().split("|")[1].strip())

        # 调整列宽
        for col in range(1, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

        wb.save(filepath)
        messagebox.showinfo("成功", f"报告已导出到:\n{filepath}")
        self.status_var.set(f"Excel报告已导出 | {filepath}")

    def create_settings_tab(self):
        frame = ttk.LabelFrame(self.tab4, text="高级参数与预设管理", padding=20)
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        ttkb.Label(frame, text="平台费率预设管理（可自定义）", font=("Microsoft YaHei", 12, "bold")).pack(anchor=W, pady=10)

        # 预设表格
        preset_frame = ttk.Frame(frame)
        preset_frame.pack(fill=X, pady=10)

        self.preset_tree = ttk.Treeview(preset_frame, columns=("平台", "佣金%", "税率%", "其他%"), show="headings", height=6)
        for col in ("平台", "佣金%", "税率%", "其他%"):
            self.preset_tree.heading(col, text=col)
            self.preset_tree.column(col, width=120, anchor=CENTER)
        self.preset_tree.pack(side=LEFT, fill=X, expand=True)

        for name, rates in PLATFORM_PRESETS.items():
            self.preset_tree.insert("", "end", values=(name, rates["commission"], rates["tax"], rates["other"]))

        # 编辑区
        edit_frame = ttk.Frame(frame)
        edit_frame.pack(fill=X, pady=15)

        ttkb.Label(edit_frame, text="自定义平台名称:").grid(row=0, column=0, sticky=W)
        self.custom_name = ttkb.Entry(edit_frame, width=15)
        self.custom_name.grid(row=0, column=1, padx=5)

        ttkb.Label(edit_frame, text="佣金率 %:").grid(row=0, column=2, sticky=W, padx=10)
        self.custom_comm = ttkb.Entry(edit_frame, width=8)
        self.custom_comm.grid(row=0, column=3)

        ttkb.Label(edit_frame, text="税率 %:").grid(row=0, column=4, sticky=W, padx=10)
        self.custom_tax = ttkb.Entry(edit_frame, width=8)
        self.custom_tax.grid(row=0, column=5)

        ttkb.Label(edit_frame, text="其他 %:").grid(row=0, column=6, sticky=W, padx=10)
        self.custom_other = ttkb.Entry(edit_frame, width=8)
        self.custom_other.grid(row=0, column=7)

        ttkb.Button(edit_frame, text="💾 保存自定义预设", command=self.save_custom_preset, bootstyle="success").grid(row=1, column=0, columnspan=4, pady=10)

        # 其他设置
        ttk.Separator(frame, orient=HORIZONTAL).pack(fill=X, pady=15)
        ttkb.Label(frame, text="其他功能设置", font=("Microsoft YaHei", 11, "bold")).pack(anchor=W)

        self.auto_save_var = tk.BooleanVar(value=True)
        ttkb.Checkbutton(frame, text="自动保存历史记录", variable=self.auto_save_var, bootstyle="round-toggle").pack(anchor=W, pady=5)

        ttkb.Button(frame, text="📁 打开历史记录文件夹", command=self.open_history_folder, bootstyle="secondary").pack(anchor=W, pady=5)

        ttkb.Button(frame, text="🔄 恢复默认设置", command=self.load_defaults, bootstyle="warning").pack(anchor=W, pady=5)

    def save_custom_preset(self):
        name = self.custom_name.get().strip()
        if not name:
            messagebox.showwarning("提示", "请输入平台名称")
            return
        try:
            comm = float(self.custom_comm.get())
            tax = float(self.custom_tax.get())
            other = float(self.custom_other.get())
            PLATFORM_PRESETS[name] = {"commission": comm, "tax": tax, "other": other}
            self.preset_tree.insert("", "end", values=(name, comm, tax, other))
            
            # 刷新两个计算页面的下拉框（解决自定义后不同步的问题）
            if hasattr(self, 'platform_combo'):
                self.platform_combo['values'] = list(PLATFORM_PRESETS.keys())
            if hasattr(self, 'roi_combo'):
                self.roi_combo['values'] = list(PLATFORM_PRESETS.keys())
            
            messagebox.showinfo("成功", f"已添加自定义平台: {name}\n\n其他页面下拉框已自动刷新！")
        except:
            messagebox.showerror("错误", "请输入正确的数值")

    def open_history_folder(self):
        path = os.path.join(os.path.expanduser("~"), ".roi_calculator_history")
        os.makedirs(path, exist_ok=True)
        messagebox.showinfo("提示", f"历史记录保存在:\n{path}\n（当前版本为内存记录）")

    def toggle_theme(self):
        current = self.style.theme.name
        new_theme = "darkly" if current != "darkly" else "cosmo"
        self.style.theme_use(new_theme)
        self.status_var.set(f"主题已切换为: {new_theme}")

    def show_help(self):
        help_text = """📖 使用说明

1. 保本价计算：输入成本、目标利润、销量、固定费用，选择平台后计算保本售价。
2. ROI计算：输入销售额、成本、推广费等，快速分析推广是否盈利。
3. 批量计算：添加多款商品，一键计算对比，导出专业Excel报告。
4. 设置：可自定义平台费率（淘宝、拼多多、抖音等主流平台已预设）。

💡 提示：
- 所有计算均实时更新
- 支持导出带格式的Excel报告
- 建议先在“高级设置”中确认当前平台费率

如有疑问，欢迎反馈！"""
        messagebox.showinfo("使用帮助", help_text)

    def add_to_history(self, calc_type, detail):
        self.history.append({
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type": calc_type,
            "detail": detail
        })
        if len(self.history) > 50:
            self.history.pop(0)

    def load_defaults(self):
        # 初始化平台
        if hasattr(self, 'platform_var'):
            self.platform_var.set("淘宝/天猫")
            self.update_platform_rates()
        if hasattr(self, 'roi_platform_var'):
            self.roi_platform_var.set("淘宝/天猫")
            self.update_roi_rates()

        self.status_var.set("默认设置已加载 | 推荐使用淘宝/天猫预设")

if __name__ == "__main__":
    root = ttkb.Window(themename="cosmo")
    app = ROICalculatorApp(root)
    root.mainloop()
